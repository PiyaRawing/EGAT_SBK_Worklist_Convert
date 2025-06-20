import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.styles import Color
from openpyxl.comments import Comment
import os
import re
import itertools
import sys

# Global variable for worklist file path
worklist_file_path = None
# Global variable for lookup data (from Respone - Do not Delete.xlsx)
response_lookup_data = {
    'AB_lookup': {}, # For J to K (Col A: Col B in Response file)
    'CD_lookup': {}  # For S to L (Col C: Col D in Response file)
}

# Global variable for the selected sheet name
selected_sheet_name = None

# Global variable for the sheet selection UI elements
sheet_selection_frame = None
sheet_option_menu = None

# Global variable for the highlighting option state
enable_highlight_var = None
# Global variable for the option to include (not skip) rows with strikethrough
include_strikethrough_rows_var = None
# Global variable for the sorting option state
enable_sort_var = None # For sorting D, E, K

# Global variable to store template rows data (including styles and merged cells)
template_rows_data = []

# Global variable for the status label
status_label = None

def get_resource_path(relative_path):
    """
    Get the absolute path to resource, works for dev and for PyInstaller.
    """
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def select_excel_file():
    """
    Opens a file dialog for the user to select an Excel file (Worklist File)
    and displays the selected file path. Then, it populates a dropdown
    with sheet names for the user to choose from.
    """
    global worklist_file_path, selected_sheet_name, sheet_selection_frame, sheet_option_menu

    new_worklist_file_path = filedialog.askopenfilename(
        title="เลือกไฟล์ Worklist Excel",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if new_worklist_file_path:
        worklist_file_path = new_worklist_file_path

        file_path_label.config(text=f"ไฟล์ Worklist: {os.path.basename(worklist_file_path)}")

        try:
            temp_workbook = openpyxl.load_workbook(worklist_file_path)
            sheet_names = temp_workbook.sheetnames
            
            active_sheet_title = None
            try:
                active_sheet_title = temp_workbook.active.title
            except Exception:
                pass

            temp_workbook.close()

            if not sheet_names:
                messagebox.showwarning("คำเตือน", "ไฟล์ Excel ที่เลือกไม่มีชีท!")
                file_path_label.config(text="ยังไม่ได้เลือกไฟล์ Worklist")
                convert_button.config(state=tk.DISABLED)
                if sheet_selection_frame:
                    sheet_selection_frame.destroy()
                    sheet_selection_frame = None
                return

            if sheet_selection_frame:
                sheet_selection_frame.destroy()

            sheet_selection_frame = tk.LabelFrame(root, text="2. เลือกชีท", padx=10, pady=10)
            sheet_selection_frame.pack(pady=5, padx=20, fill="x")

            sheet_label = tk.Label(sheet_selection_frame, text="เลือกชีท:")
            sheet_label.pack(side=tk.LEFT, padx=(0, 10))

            selected_sheet_name = tk.StringVar(root)
            if active_sheet_title and active_sheet_title in sheet_names:
                selected_sheet_name.set(active_sheet_title)
            else:
                selected_sheet_name.set(sheet_names[0])

            sheet_option_menu = tk.OptionMenu(sheet_selection_frame, selected_sheet_name, *sheet_names)
            sheet_option_menu.pack(side=tk.LEFT, fill="x", expand=True)
            
            convert_button.config(state=tk.NORMAL)

        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถอ่านไฟล์ Excel ได้: {e}\n"
                                               f"โปรดตรวจสอบว่าไฟล์ไม่ได้ถูกเปิดอยู่หรือเสียหาย")
            file_path_label.config(text="ยังไม่ได้เลือกไฟล์ Worklist")
            convert_button.config(state=tk.DISABLED)
            if sheet_selection_frame:
                sheet_selection_frame.destroy()
                sheet_selection_frame = None
            worklist_file_path = None

    else:
        file_path_label.config(text="ยังไม่ได้เลือกไฟล์ Worklist")
        convert_button.config(state=tk.DISABLED)
        if sheet_selection_frame:
            sheet_selection_frame.destroy()
            sheet_selection_frame = None
        worklist_file_path = None


def split_j_column_data(text_data):
    """
    Splits text data from Column J based on '.-', '. ', ',', or '/' delimiters.
    Ensures each split part ends with a dot if it represents a meaningful segment.
    Handles leading/trailing pipe characters and re-adds them to split parts.
    """
    if not isinstance(text_data, str):
        return [text_data]

    original_had_pipes = text_data.startswith('|') and text_data.endswith('|')
    clean_text = text_data.strip('|')

    delimiters = [".-", ". ", ",", "/"]
    temp_delimiter = "###SPLIT_POINT###"

    processed_text = clean_text
    for delim in delimiters:
        processed_text = processed_text.replace(delim, temp_delimiter)
    
    raw_parts = [p.strip() for p in processed_text.split(temp_delimiter) if p.strip()]

    parts = []
    for part in raw_parts:
        if part:
            if not part.endswith('.'):
                parts.append(part + '.')
            else:
                parts.append(part)
    
    if not parts:
        return [None] if clean_text.strip() else [None]

    if original_had_pipes:
        parts = [f"|{p}|" for p in parts]

    return parts


def split_i_column_data(text_data):
    """
    Splits text data from Column I based on 'number dot' pattern (e.g., '1.', '2.', '10.').
    Crucially, it does NOT split if the 'number dot' pattern is found inside parentheses.
    Handles leading/trailing pipe characters.
    Also implements Process D: Removes '/.', '/', '.', or '\' if found at the end of each part.
    """
    if not isinstance(text_data, str):
        return [text_data]

    original_had_pipes = text_data.startswith('|') and text_data.endswith('|')
    clean_text = text_data.strip('|')

    items = []
    valid_split_indices = [0]

    paren_level = 0
    for m in re.finditer(r'\d+\.', clean_text):
        match_start = m.start()
        
        current_paren_level = 0
        for char_idx in range(match_start):
            if clean_text[char_idx] == '(':
                current_paren_level += 1
            elif clean_text[char_idx] == ')':
                current_paren_level -= 1

        if current_paren_level == 0:
            if match_start not in valid_split_indices:
                valid_split_indices.append(match_start)
    
    valid_split_indices = sorted(list(set(valid_split_indices)))
    
    for k in range(len(valid_split_indices)):
        start_idx = valid_split_indices[k]
        end_idx = valid_split_indices[k+1] if k+1 < len(valid_split_indices) else len(clean_text)
        
        segment = clean_text[start_idx:end_idx].strip()
        if segment:
            items.append(segment)
    
    if not items:
        return [None]

    processed_parts = []
    for part in items:
        if isinstance(part, str):
            temp_part = part.strip()

            if temp_part.endswith('/.'):
                temp_part = temp_part[:-2].strip()

            temp_part = temp_part.rstrip('/.\\') 

            processed_parts.append(temp_part)
        else:
            processed_parts.append(part)
    
    if not processed_parts:
        return [None]

    if original_had_pipes:
        processed_parts = [f"|{p}|" for p in processed_parts]

    return processed_parts

def load_lookup_data():
    """
    Loads data from 'Respone - Do not Delete.xlsx' into a dictionary for VLOOKUP.
    Keys are from Column A and C, values are from Column B and D respectively.
    Specifically loads from the 'Respone' sheet.
    """
    global response_lookup_data
    response_lookup_data = {
        'AB_lookup': {},
        'CD_lookup': {}
    }

    response_file_path = get_resource_path("Respone - Do not Delete.xlsx")

    if not os.path.exists(response_file_path):
        messagebox.showerror(
            "ข้อผิดพลาด",
            f"ไม่พบไฟล์ VLOOKUP: '{response_file_path}'\n"
            "กรุณาตรวจสอบว่าไฟล์ 'Respone - Do not Delete.xlsx' อยู่ในโฟลเดอร์เดียวกับโปรแกรม"
        )
        return False

    try:
        lookup_workbook = openpyxl.load_workbook(response_file_path)
        if 'Respone' in lookup_workbook.sheetnames:
            lookup_sheet = lookup_workbook['Respone']
        else:
            messagebox.showerror(
                "ข้อผิดพลาดชีท",
                f"ไม่พบชีท 'Respone' ในไฟล์ '{response_file_path}'\n"
                "โปรดตรวจสอบชื่อชีทในไฟล์ Respone - Do not Delete.xlsx"
            )
            return False

        for row_idx in range(1, lookup_sheet.max_row + 1):
            key_ab = lookup_sheet.cell(row=row_idx, column=1).value
            value_ab = lookup_sheet.cell(row=row_idx, column=2).value
            if key_ab is not None:
                response_lookup_data['AB_lookup'][str(key_ab).strip()] = value_ab

            key_cd = lookup_sheet.cell(row=row_idx, column=3).value
            value_cd = lookup_sheet.cell(row=row_idx, column=4).value
            if key_cd is not None:
                response_lookup_data['CD_lookup'][str(key_cd).strip()] = value_cd
        return True
    except Exception as e:
        messagebox.showerror(
            "ข้อผิดพลาดในการโหลดไฟล์ VLOOKUP",
            f"เกิดข้อผิดพลาดขณะโหลดไฟล์ 'Respone - Do not Delete.xlsx': {e}"
        )
        return False

def load_template_rows():
    """
    Loads the first two rows (including values, styles, comments, and merged cells)
    from the 'Template' sheet of 'Respone - Do not Delete.xlsx'.
    """
    global template_rows_data
    template_rows_data = []

    response_file_path = get_resource_path("Respone - Do not Delete.xlsx")

    if not os.path.exists(response_file_path):
        messagebox.showerror(
            "ข้อผิดพลาด",
            f"ไม่พบไฟล์ Template: '{response_file_path}'\n"
            "กรุณาตรวจสอบว่าไฟล์ 'Respone - Do not Delete.xlsx' อยู่ในโฟลเดอร์เดียวกับโปรแกรม"
        )
        return False
    
    try:
        template_workbook = openpyxl.load_workbook(response_file_path)
        if 'Template' in template_workbook.sheetnames:
            template_sheet = template_workbook['Template']
        else:
            messagebox.showerror(
                "ข้อผิดพลาดชีท",
                f"ไม่พบชีท 'Template' ในไฟล์ '{response_file_path}'\n"
                "โปรดตรวจสอบชื่อชีทในไฟล์ Respone - Do not Delete.xlsx"
            )
            return False

        merged_cells_ranges_from_template = []
        for merged_range in template_sheet.merged_cells.ranges:
            merged_cells_ranges_from_template.append(str(merged_range)) 
        
        template_rows_data.append(merged_cells_ranges_from_template)

        for r_idx in range(1, 3):
            row_cells_data = {}
            for c_idx in range(1, template_sheet.max_column + 1): 
                cell = template_sheet.cell(row=r_idx, column=c_idx)
                
                cell_data = {
                    'value': cell.value,
                    'fill': PatternFill(start_color=cell.fill.start_color, 
                                        end_color=cell.fill.end_color, 
                                        fill_type=cell.fill.fill_type) if cell.fill else None, 
                    'font': cell.font.copy() if cell.font else None,
                    'border': cell.border.copy() if cell.border else None,
                    'alignment': cell.alignment.copy() if cell.alignment else None,
                    'number_format': cell.number_format,
                    'comment': Comment(cell.comment.text, cell.comment.author) if cell.comment else None 
                }
                row_cells_data[cell.column_letter] = cell_data
            template_rows_data.append(row_cells_data)
        
        template_workbook.close()
        return True
    except Exception as e:
        messagebox.showerror(
            "ข้อผิดพลาดในการโหลด Template",
            f"เกิดข้อผิดพลาดขณะโหลดชีท 'Template' จากไฟล์ 'Respone - Do not Delete.xlsx': {e}"
        )
        return False

def run_conversion_process():
    """
    Wrapper function to handle UI state before and after conversion.
    """
    global status_label

    convert_button.config(state=tk.DISABLED)
    status_label.config(text="กำลังแปลงข้อมูล... โปรดรอสักครู่")
    root.update_idletasks()

    try:
        convert_to_maximo()
    finally:
        convert_button.config(state=tk.NORMAL)
        root.update_idletasks()


def convert_to_maximo():
    """
    Reads data from specified columns (B, F, H, I, J) starting from Row 3 of the
    selected Worklist file and writes them to new columns (D, E, I, J, K, S) in a new Excel file.
    Column J data is split and duplicated across new rows as necessary.
    Column I data is split (excluding within parentheses and with '/.' removed from end) and duplicated across new rows.
    Process E (VLOOKUP) is applied to Column J to populate Column K, and to Column S to populate Column S (overwrite).
    Process F (Counting in G and H based on D, E, J groups) is applied.
    Process for highlighting cells in Column I (if enabled and condition met).
    Process G (Copying template rows) is applied at the beginning of the new sheet.
    The user is prompted to choose the save location for the new file.
    
    New Feature: Option to include/exclude rows with strikethrough formatting.
    New Feature: Option to sort by Column D, then Column E, then Column K.
    """
    global selected_sheet_name
    global enable_highlight_var
    global include_strikethrough_rows_var
    global enable_sort_var
    global status_label

    if not worklist_file_path:
        messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ Worklist ก่อนดำเนินการ!")
        status_label.config(text="โปรดเลือกไฟล์ Worklist")
        return
    
    if not selected_sheet_name or not selected_sheet_name.get():
        messagebox.showwarning("คำเตือน", "กรุณาเลือกชีทที่จะแปลงข้อมูล!")
        status_label.config(text="โปรดเลือกชีท")
        return

    if not load_lookup_data():
        status_label.config(text="ข้อผิดพลาดในการโหลด VLOOKUP Data")
        return

    if not load_template_rows():
        status_label.config(text="ข้อผิดพลาดในการโหลด Template Data")
        return

    try:
        source_workbook = openpyxl.load_workbook(worklist_file_path)
        source_sheet = source_workbook[selected_sheet_name.get()]

        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Template"

        # --- Process G: Copy template rows to the new sheet ---
        if template_rows_data and isinstance(template_rows_data[0], list):
            for merged_range_str in template_rows_data[0]:
                try:
                    new_sheet.merge_cells(merged_range_str)
                except Exception as merge_err:
                    print(f"Warning: Could not merge cells {merged_range_str}: {merge_err}") 
        
        for r_idx_template_data in range(1, 3):
            row_data_to_copy = template_rows_data[r_idx_template_data] 
            for col_letter, cell_info in row_data_to_copy.items():
                target_cell = new_sheet[f"{col_letter}{r_idx_template_data}"]
                target_cell.value = cell_info['value']
                
                if cell_info['fill']:
                    target_cell.fill = cell_info['fill']
                if cell_info['font']:
                    target_cell.font = cell_info['font']
                if cell_info['border']:
                    target_cell.border = cell_info['border']
                if cell_info['alignment']:
                    target_cell.alignment = cell_info['alignment']
                if cell_info['number_format']:
                    target_cell.number_format = cell_info['number_format']
                if cell_info['comment']:
                    target_cell.comment = cell_info['comment']

        highlight_fill = PatternFill(start_color='FF9E00', end_color='FF9E00', fill_type='solid')
        
        # This list will store all processed row data before writing to new_sheet
        # Format: [ (D_value, E_value, K_value, I_value, J_value, S_value, Original_I_Worklist), ... ]
        # K_value is added to the tuple for sorting.
        all_processed_rows_data = [] 
        
        for source_row_idx in range(5, source_sheet.max_row + 1):
            status_label.config(text=f"กำลังแปลงข้อมูล... (กำลังประมวลผลแถวที่ {source_row_idx})")
            root.update_idletasks()

            if not include_strikethrough_rows_var.get(): 
                skip_row_due_to_strikethrough = False
                for col_idx in range(1, source_sheet.max_column + 1):
                    cell = source_sheet.cell(row=source_row_idx, column=col_idx)
                    if cell.font and cell.font.strike:
                        skip_row_due_to_strikethrough = True
                        break
                
                if skip_row_due_to_strikethrough:
                    continue

            data_col_b = source_sheet.cell(row=source_row_idx, column=2).value
            data_col_i = source_sheet.cell(row=source_row_idx, column=9).value
            data_col_f = source_sheet.cell(row=source_row_idx, column=6).value
            data_col_h_for_i = source_sheet.cell(row=source_row_idx, column=8).value
            data_col_j = source_sheet.cell(row=source_row_idx, column=10).value

            data_col_i_worklist_for_s = source_sheet.cell(row=source_row_idx, column=9).value

            j_parts = split_j_column_data(data_col_j)
            i_parts = split_i_column_data(data_col_h_for_i)

            for j_part_item in j_parts:
                lookup_key_k = str(j_part_item).strip() if j_part_item is not None else ""
                vlookup_result_k = response_lookup_data['AB_lookup'].get(lookup_key_k, None)

                lookup_key_s = str(data_col_i_worklist_for_s).strip() if data_col_i_worklist_for_s is not None else ""
                vlookup_result_s = response_lookup_data['CD_lookup'].get(lookup_key_s, None)
                
                final_s_value = vlookup_result_s if vlookup_result_s is not None else data_col_i_worklist_for_s


                for i_part_item in i_parts:
                    all_processed_rows_data.append((
                        data_col_b,       # Index 0: D_value
                        data_col_f,       # Index 1: E_value
                        vlookup_result_k, # Index 2: K_value (NEW for sorting)
                        i_part_item,      # Index 3: I_value
                        j_part_item,      # Index 4: J_value
                        final_s_value,    # Index 5: S_value
                        data_col_i        # Index 6: Original Col I from Worklist for output Col V
                    ))
        
        # --- Process H: Conditional Sorting by Column D then Column E then Column K ---
        if enable_sort_var.get():
            # Sort the collected data: primary key is D (index 0), secondary E (index 1), tertiary K (index 2)
            # Use a lambda function for robust string conversion for sorting (handles None and numbers gracefully)
            all_processed_rows_data.sort(key=lambda x: (
                str(x[0] or ''),  # Column D
                str(x[1] or ''),  # Column E
                str(x[2] or '')   # Column K (NEW)
            ))
            
        rows_for_f_process = [] # Will store (D, E, J) and current_output_row_idx for Process F
        current_output_row_idx = 3 # Start writing from row 3

        for row_data in all_processed_rows_data:
            # Unpack the data based on the order it was stored in all_processed_rows_data
            data_col_d = row_data[0]
            data_col_e = row_data[1]
            data_col_k_output = row_data[2] # Use the K value directly from the stored data
            data_col_i_output = row_data[3]
            data_col_j_output = row_data[4]
            data_col_s_output = row_data[5]
            data_col_v_output = row_data[6]

            new_sheet.cell(row=current_output_row_idx, column=4).value = data_col_d # Col D
            new_sheet.cell(row=current_output_row_idx, column=5).value = data_col_e # Col E
            
            cell_i_output_target = new_sheet.cell(row=current_output_row_idx, column=9)
            cell_i_output_target.value = data_col_i_output # Col I

            new_sheet.cell(row=current_output_row_idx, column=10).value = data_col_j_output # Col J
            new_sheet.cell(row=current_output_row_idx, column=11).value = data_col_k_output # Col K (Now written from sorted data)
            new_sheet.cell(row=current_output_row_idx, column=19).value = data_col_s_output # Col S
            new_sheet.cell(row=current_output_row_idx, column=22).value = data_col_v_output # Col V

            if enable_highlight_var.get():
                if isinstance(cell_i_output_target.value, str) and len(cell_i_output_target.value) > 85:
                    cell_i_output_target.fill = highlight_fill

            # Store (D, E, J) values and their corresponding row index in the new sheet
            # It's important that this list is created *after* the main sorting
            # and contains the row index in the *newly written* sheet.
            rows_for_f_process.append((
                (str(data_col_d) if data_col_d is not None else "",
                 str(data_col_e) if data_col_e is not None else "",
                 str(data_col_j_output) if data_col_j_output is not None else ""),
                current_output_row_idx
            ))
            
            current_output_row_idx += 1


        # --- Process F: Counting in Column G and H based on D, E, J groups ---
        
        # !!! IMPORTANT FIX !!!
        # Remove the internal sort here. The main sorting for D, E, K
        # has already been applied to all_processed_rows_data.
        # For groupby to work correctly with D, E, J, the data
        # MUST already be sorted by D, E, J. If D, E, K is the desired main sort,
        # and D, E, J is a sub-grouping, the previous sort ensures this.
        # If the user wants a D,E,K sort primarily, and then grouping by D,E,J,
        # this still works because D,E,K sort will naturally group D,E,J together.
        # If the sort key for groupby is different from the sort key of the main data,
        # you might need to sort again here, but it would override the D,E,K sort.
        # Given the request is D then E then K, removing this sort is the correct approach.
        # rows_for_f_process.sort(key=lambda x: x[0]) # <--- REMOVED THIS LINE

        # To ensure groupby works as expected, it's safer to sort by D, E, J here,
        # but this will override the D, E, K sort.
        # A better approach for combined sorting and grouping is to ensure the primary sort
        # encompasses the grouping keys, or to perform the grouping before the final sort if order matters.

        # Let's re-evaluate. If we want D, E, K as the final order, but still need to count G/H based on D, E, J.
        # The current approach (sorting all_processed_rows_data by D, E, K then iterating to write)
        # means that `rows_for_f_process` will already be in an order that has D, E, J grouped IF
        # K values within a D, E group are also naturally grouped by J.
        # If D, E, J is the ONLY basis for counting, and the D, E, K sort
        # could break the contiguity of J within D,E, then a separate sort for `rows_for_f_process` IS needed.

        # Let's consider the user's intent: "sort from D then E then K". This implies the final visual order.
        # The counting for G and H (Process F) is based on groups of D, E, J.
        # If we sort D, E, K, then write, the D, E, J groups might not be contiguous anymore for `itertools.groupby`.
        # To fix this, we need to apply Process F *before* the final D, E, K sort, or handle the counting
        # in a way that respects the final sort order.

        # A more robust approach: Calculate counts *before* the final sort, then store them.
        # Let's revert to storing the counts in the `all_processed_rows_data` for a cleaner flow.

        # Let's collect items for grouping (D, E, J values) and their original index to update counts later.
        # This will be done before the final D, E, K sort.
        
        # --- NEW STRATEGY FOR PROCESS F (Counting G & H) ---
        # 1. Group data by (D, E, J) to calculate counts.
        # 2. Store the counts with the corresponding (D, E, J) group.
        # 3. When writing, retrieve the correct count for the D, E, J of that row.

        # We need to map (D, E, J) to a count
        group_counts = {} # Key: (D_value, E_value, J_value), Value: starting count
        
        # Create a temporary list to hold data for grouping, maintaining original order
        # This list will be sorted ONLY for the purpose of itertools.groupby
        temp_data_for_grouping = []
        for i, row_data in enumerate(all_processed_rows_data):
            # Extract D, E, J which are at indices 0, 1, 4 in all_processed_rows_data
            d_val = str(row_data[0] or '')
            e_val = str(row_data[1] or '')
            j_val = str(row_data[4] or '')
            temp_data_for_grouping.append(((d_val, e_val, j_val), i)) # Store (D,E,J) key and original index

        # Sort this temporary list by the D, E, J keys for groupby
        temp_data_for_grouping.sort(key=lambda x: x[0])

        # Apply grouping and calculate counts
        for group_key, group_iter in itertools.groupby(temp_data_for_grouping, key=lambda x: x[0]):
            current_count = 10
            for _, original_index in group_iter:
                # Update the all_processed_rows_data list at the original index
                # with the calculated G and H values.
                # We need to expand all_processed_rows_data to hold G and H values.
                # Let's re-structure `all_processed_rows_data` to be a dictionary or object if it gets too complex.
                # For now, let's assume we can extend the tuple in all_processed_rows_data.
                # This requires that `all_processed_rows_data` is a list of lists/mutable objects initially.
                # Or, we can store the calculated G and H values in a separate map and look them up during write.
                
                # Let's use a separate map to store calculated G/H values, indexed by the original tuple position.
                # This ensures `all_processed_rows_data` remains sortable as a tuple.
                if 'calculated_g_h' not in globals(): # Initialize if not exists
                    global calculated_g_h
                    calculated_g_h = {} # Key: original_index, Value: (G_value, H_value)
                
                calculated_g_h[original_index] = (current_count, current_count)
                current_count += 10

        # --- Re-apply the D, E, K sort to all_processed_rows_data ---
        # This sort should be the *last* major sort before writing.
        if enable_sort_var.get():
            all_processed_rows_data.sort(key=lambda x: (
                str(x[0] or ''),  # Column D
                str(x[1] or ''),  # Column E
                str(x[2] or '')   # Column K
            ))

        # Now, write the sorted data and apply the pre-calculated G/H values
        current_output_row_idx = 3 # Start writing from row 3
        # We need a way to link back to the original index from `all_processed_rows_data`
        # as it was *before* the D, E, K sort to retrieve the G/H values.
        # This means `all_processed_rows_data` itself needs to contain its original index.

        # Let's restart the loop with a clearer structure:
        # Step 1: Collect all base data (D, E, K, I, J, S, V) along with its *original_source_row_idx*
        # Step 2: Perform the D,E,J grouping logic on a temporary structure that includes original_source_row_idx
        #         to calculate and store G,H values associated with each original_source_row_idx.
        # Step 3: Add the calculated G,H values to the main `all_processed_rows_data` tuples.
        # Step 4: Perform the final D,E,K sort on `all_processed_rows_data`.
        # Step 5: Write the data to the new sheet.

        # RE-STRUCTURING `all_processed_rows_data` COLLECTION
        all_processed_rows_data = [] # Reset this list
        # Store: (D_value, E_value, K_value, I_value, J_value, S_value, Original_I_Worklist, temporary_placeholder_G, temporary_placeholder_H)
        # Use placeholder for G and H initially (e.g., None)
        
        # We need a mapping from a unique ID per generated row to its G/H values
        # Let's just calculate it directly when the row is generated and store it.
        # This makes the D,E,K sort potentially break the G/H grouping.
        # The best way is to calculate G,H for (D,E,J) groups first, then assign them to the rows.

        # Let's revert to a simpler method by calculating G and H after sorting by D, E, K.
        # This means G and H will be based on the D, E, J of the rows *in the final sorted order*.
        # If this is acceptable (i.e., G/H can reset in the middle of a D,E group if K causes it), then this is simpler.
        # However, typically G/H (sequence numbers) should be based on consistent grouping keys.

        # Let's assume the user wants G/H to be contiguous based on D, E, J regardless of K's influence.
        # This implies that the D, E, J grouping/counting must happen *before* the final D, E, K sort.
        # This means `all_processed_rows_data` needs to store G and H values.

        # REVISED `all_processed_rows_data` Structure for better flow:
        # all_processed_rows_data will contain dictionaries for flexibility:
        # { 'D': val, 'E': val, 'K': val, 'I': val, 'J': val, 'S': val, 'V': val, 'G': val, 'H': val, 'original_row_idx': some_idx }
        
        rows_to_process_and_sort = [] # This will hold dictionaries
        
        # First Pass: Extract data, apply initial lookups and splits, prepare for grouping
        # Store original source row index to help with debugging/tracking.
        for source_row_idx in range(5, source_sheet.max_row + 1):
            status_label.config(text=f"กำลังแปลงข้อมูล... (กำลังประมวลผลแถวที่ {source_row_idx})")
            root.update_idletasks()

            if not include_strikethrough_rows_var.get(): 
                skip_row_due_to_strikethrough = False
                for col_idx in range(1, source_sheet.max_column + 1):
                    cell = source_sheet.cell(row=source_row_idx, column=col_idx)
                    if cell.font and cell.font.strike:
                        skip_row_due_to_strikethrough = True
                        break
                if skip_row_due_to_strikethrough:
                    continue

            data_col_b = source_sheet.cell(row=source_row_idx, column=2).value
            data_col_i = source_sheet.cell(row=source_row_idx, column=9).value # Original Column I from Worklist (for Col V)
            data_col_f = source_sheet.cell(row=source_row_idx, column=6).value
            data_col_h_for_i = source_sheet.cell(row=source_row_idx, column=8).value # Original Column H for Col I after splitting
            data_col_j = source_sheet.cell(row=source_row_idx, column=10).value

            data_col_i_worklist_for_s = source_sheet.cell(row=source_row_idx, column=9).value

            j_parts = split_j_column_data(data_col_j)
            i_parts = split_i_column_data(data_col_h_for_i)

            for j_part_item in j_parts:
                lookup_key_k = str(j_part_item).strip() if j_part_item is not None else ""
                vlookup_result_k = response_lookup_data['AB_lookup'].get(lookup_key_k, None)

                lookup_key_s = str(data_col_i_worklist_for_s).strip() if data_col_i_worklist_for_s is not None else ""
                vlookup_result_s = response_lookup_data['CD_lookup'].get(lookup_key_s, None)
                final_s_value = vlookup_result_s if vlookup_result_s is not None else data_col_i_worklist_for_s

                for i_part_item in i_parts:
                    rows_to_process_and_sort.append({
                        'D': data_col_b,
                        'E': data_col_f,
                        'I': i_part_item,
                        'J': j_part_item,
                        'K': vlookup_result_k,
                        'S': final_s_value,
                        'V': data_col_i, # Original Worklist Col I
                        'G': None, # Placeholder for G
                        'H': None  # Placeholder for H
                    })

        # Second Pass: Calculate G and H counts based on (D, E, J) groups
        # Create a copy and sort by D, E, J to enable itertools.groupby
        temp_for_grouping_gh = sorted(rows_to_process_and_sort, key=lambda x: (
            str(x['D'] or ''), 
            str(x['E'] or ''), 
            str(x['J'] or '')
        ))
        
        # Iterate through grouped data and assign G/H values
        for group_key, group_iter in itertools.groupby(temp_for_grouping_gh, key=lambda x: (
            str(x['D'] or ''), 
            str(x['E'] or ''), 
            str(x['J'] or '')
        )):
            current_count = 10
            for row_dict in group_iter:
                # Find the original row in rows_to_process_and_sort and update its G/H
                # This is inefficient for large datasets, but simple.
                # A better way for large data would be to map back using a unique ID or process in place.
                # For typical Excel sizes (few thousands rows), this should be acceptable.
                for original_row_dict in rows_to_process_and_sort:
                    if original_row_dict is row_dict: # Check if it's the exact same dictionary object
                        original_row_dict['G'] = current_count
                        original_row_dict['H'] = current_count
                        break
                current_count += 10
        
        # Third Pass: Perform the final sort (D, E, K) if enabled
        if enable_sort_var.get():
            rows_to_write = sorted(rows_to_process_and_sort, key=lambda x: (
                str(x['D'] or ''),
                str(x['E'] or ''),
                str(x['K'] or '')
            ))
        else:
            rows_to_write = rows_to_process_and_sort # No sorting, use original order of generation

        # Fourth Pass: Write the data to the new sheet
        current_output_row_idx = 3
        for row_data in rows_to_write:
            new_sheet.cell(row=current_output_row_idx, column=4).value = row_data['D']
            new_sheet.cell(row=current_output_row_idx, column=5).value = row_data['E']
            new_sheet.cell(row=current_output_row_idx, column=7).value = row_data['G'] # G from calculated
            new_sheet.cell(row=current_output_row_idx, column=8).value = row_data['H'] # H from calculated
            
            cell_i_output_target = new_sheet.cell(row=current_output_row_idx, column=9)
            cell_i_output_target.value = row_data['I'] # I

            new_sheet.cell(row=current_output_row_idx, column=10).value = row_data['J']
            new_sheet.cell(row=current_output_row_idx, column=11).value = row_data['K']
            new_sheet.cell(row=current_output_row_idx, column=19).value = row_data['S']
            new_sheet.cell(row=current_output_row_idx, column=22).value = row_data['V']

            if enable_highlight_var.get():
                if isinstance(cell_i_output_target.value, str) and len(cell_i_output_target.value) > 85:
                    cell_i_output_target.fill = highlight_fill
            
            current_output_row_idx += 1

        default_file_name = f"{selected_sheet_name.get()} converted.xlsx"

        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="บันทึกไฟล์ Maximo Data เป็น",
            initialfile=default_file_name
        )

        if not output_file_path:
            messagebox.showinfo("ยกเลิก", "การบันทึกไฟล์ถูกยกเลิก")
            status_label.config(text="การแปลงข้อมูลถูกยกเลิก")
            return

        new_workbook.save(output_file_path)

        messagebox.showinfo(
            "สำเร็จ",
            f"การแปลงข้อมูลเสร็จสมบูรณ์ และบันทึกเป็นไฟล์ใหม่แล้วที่:\n{output_file_path}"
        )
        status_label.config(text="การแปลงข้อมูลเสร็จสมบูรณ์")

    except FileNotFoundError:
        messagebox.showerror("ข้อผิดพลาด", "ไม่พบไฟล์ Worklist ที่ระบุ กรุณาตรวจสอบเส้นทางไฟล์")
        status_label.config(text="เกิดข้อผิดพลาด: ไม่พบไฟล์")
    except Exception as e:
        messagebox.showerror("เกิดข้อผิดพลาด", f"เกิดข้อผิดพลาดในการประมวลผลไฟล์: {e}\n"
                                               f"โปรดตรวจสอบว่าไฟล์ Excel ไม่ได้ถูกเปิดอยู่")
        status_label.config(text=f"เกิดข้อผิดพลาด: {e}")

# --- GUI Setup ---
root = tk.Tk()
root.title("Excel Worklist Converter")
root.geometry("720x550")
root.resizable(False, False)
root.iconbitmap("./transfer.ico")
root.option_add("*font", "Tahoma 14")

worklist_file_path = None
response_lookup_data = {
    'AB_lookup': {},
    'CD_lookup': {}
}
selected_sheet_name = None 
sheet_selection_frame = None
sheet_option_menu = None

enable_highlight_var = tk.BooleanVar(value=False)
include_strikethrough_rows_var = tk.BooleanVar(value=False) 
enable_sort_var = tk.BooleanVar(value=False)

template_rows_data = []

file_frame = tk.LabelFrame(root, text="1. Worklist File", padx=10, pady=10)
file_frame.pack(pady=10, padx=20, fill="x")

select_file_button = tk.Button(file_frame, text="เลือกไฟล์ Excel", command=select_excel_file)
select_file_button.pack(side=tk.LEFT, padx=(0, 10))

file_path_label = tk.Label(file_frame, text="ยังไม่ได้เลือกไฟล์ Worklist", wraplength=350, justify="left")
file_path_label.pack(side=tk.LEFT, fill="x", expand=True)

options_frame = tk.LabelFrame(root, text="ตัวเลือกผลลัพธ์", padx=10, pady=5)
options_frame.pack(pady=5, padx=20, fill="x")

highlight_check = tk.Checkbutton(
    options_frame, 
    text="1. เปิดใช้งานการเน้นสี (Column Activity หากตัวอักษรเกิน 85 ตัว)",
    variable=enable_highlight_var,
    bg="orange"
)
highlight_check.pack(side=tk.TOP, anchor="w", pady=2) 

strikethrough_check = tk.Checkbutton(
    options_frame,
    text="2. ดึง Row ที่มี strikethrough (ไม่ดึงเป็นค่าเริ่มต้น)",
    variable=include_strikethrough_rows_var
)
strikethrough_check.pack(side=tk.TOP, anchor="w", pady=2)

sort_check = tk.Checkbutton(
    options_frame,
    text="3. เรียงลำดับข้อมูล (จำเป็นถ้าไม่เรียง TASK ORDER อาจจะผิดได้)", # Updated text
    variable=enable_sort_var
)
sort_check.pack(side=tk.TOP, anchor="w", pady=2)

convert_button = tk.Button(root, text="Convert to Maximo", command=run_conversion_process, state=tk.DISABLED)
convert_button.pack(pady=20)

status_label = tk.Label(root, text="", fg="blue", font=("Tahoma", 12))
status_label.pack(pady=5)

update_info_label = tk.Label(
    root,
    text="โดย : นศ.ฝึกงาน ปิยะ  ระวิงทอง | อัปเดตล่าสุด : 21/6/2568",
    font=("Tahoma", 10),
    fg="gray"
)
update_info_label.pack(side=tk.BOTTOM, anchor="se", padx=10, pady=5)

root.mainloop()